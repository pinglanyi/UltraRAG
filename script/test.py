import openpyxl
import requests
import time
import os
import glob

INPUT_DIR = "/home/wangzilong/UltraRAG/test/input"
OUTPUT_DIR = "/home/wangzilong/UltraRAG/test/output"
API_URL = "http://10.32.1.172:13229/query"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# 获取输入文件夹下所有xlsx文件
input_files = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
print(f"共找到 {len(input_files)} 个文件")

for filepath in input_files:
    filename = os.path.basename(filepath)
    output_path = os.path.join(OUTPUT_DIR, f"结果_{filename}")
    print(f"\n处理文件: {filename}")

    # 读取原始数据
    wb_in = openpyxl.load_workbook(filepath)
    ws_in = wb_in.worksheets[0]

    # 提取问题和希望的答案（第3行开始，B列=问题，C列=希望的答案）
    questions = []
    for row in ws_in.iter_rows(min_row=3, max_row=ws_in.max_row, values_only=True):
        seq, question, expected_answer = row[0], row[1], row[2]
        if question:
            questions.append((seq, question, expected_answer or ""))

    print(f"共 {len(questions)} 个问题")

    # 创建输出Excel
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "测试结果"

    # 写表头
    headers = ["序号", "问题", "希望的答案", "answer", "thinking", "耗时(秒)"]
    for col, h in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, name="Arial")

    # 逐一调用接口
    for i, (seq, question, expected) in enumerate(questions):
        print(f"[{i+1}/{len(questions)}] 提问: {question[:50]}...")
        try:
            start = time.time()
            resp = requests.post(API_URL, json={"query": question}, timeout=120)
            elapsed = round(time.time() - start, 2)
            data = resp.json()
            answer = data.get("answer", "")
            print(answer)
            thinking = data.get("thinking", "")
        except Exception as e:
            elapsed = 0
            answer = f"请求失败: {e}"
            thinking = ""

        ws_out.append([seq, question, expected, answer, thinking, elapsed])
        time.sleep(0.5)

    # 调整列宽
    ws_out.column_dimensions['A'].width = 6
    ws_out.column_dimensions['B'].width = 40
    ws_out.column_dimensions['C'].width = 50
    ws_out.column_dimensions['D'].width = 50
    ws_out.column_dimensions['E'].width = 60
    ws_out.column_dimensions['F'].width = 10

    # 设置自动换行
    for row in ws_out.iter_rows(min_row=2, max_row=ws_out.max_row):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
            cell.font = openpyxl.styles.Font(name="Arial")

    wb_out.save(output_path)
    print(f"结果已保存到: {output_path}")

print(f"\n全部完成！")